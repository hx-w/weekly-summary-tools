# -*- coding: utf-8 -*-

import socket
import re
import os
import shutil
import openpyxl
import datetime
from config import gconfig

align = openpyxl.styles.Alignment(
    horizontal='left', vertical='center', wrap_text=False)


async def __exec(source_dir: str, file_list: list, dist_file: str, template_file: str, newcol_pattern: re.Pattern = None):
    shutil.copyfile(template_file, dist_file)

    dist_book = openpyxl.load_workbook(dist_file)
    dist_sheet = dist_book['工作任务项']
    dist_sheet.delete_rows(2, dist_sheet.max_row + 1)
    col_name, col_date = 1, 0
    for idx in range(1, 1 + dist_sheet.max_column):
        if dist_sheet.cell(1, idx).value == '日期（年/月/日）':
            col_date = idx - 1
            break
    total_order = 1
    for eachfile in file_list:
        newcol = re.findall(newcol_pattern, eachfile)[
            0] if newcol_pattern else None
        full_path = os.path.join(source_dir, eachfile)
        source_sheet = openpyxl.load_workbook(
            full_path, read_only=True)['工作任务项']
        # print(source_sheet[1])
        for row_idx in range(2, source_sheet.max_row + 1):
            row_element = []
            for col_idx in range(1, source_sheet.max_column + 1):
                row_element.append(source_sheet.cell(
                    row=row_idx, column=col_idx).value)
            if None in row_element:
                break
            if newcol:
                row_element.insert(col_name, newcol)
            if row_element[col_date] and not isinstance(row_element[col_date], str):
                row_element[col_date] = datetime.datetime.strftime(
                    row_element[col_date], "%Y/%m/%d")
            row_element[0] = total_order
            dist_sheet.append(row_element)
            total_order += 1

    # align to left
    for row in dist_sheet.iter_rows(min_row=2):
        for col in row:
            col.alignment = align

    # update pivot table
    for sheet_idx in range(len(dist_book.sheetnames)):
        if dist_book.sheetnames[sheet_idx] == '工作任务项':
            continue
        try:
            pivot_sheet = dist_book[dist_book.sheetnames[sheet_idx]]
            pivot = pivot_sheet._pivots[0]  # 任何一个都可以共享同一个缓存
            boundary = f'A1:{chr(ord("A") + dist_sheet.max_column - 1)}{dist_sheet.max_row}'
            pivot.cache.cacheSource.worksheetSource.ref = boundary
            pivot.cache.refreshOnLoad = True  # 刷新加载
        except:
            pass
    dist_book.save(dist_file)


async def info_week_list(reverse: bool, single_week: bool) -> list:
    weeklist = list(filter(
        lambda x: re.match(gconfig.week_pattern, x),
        os.listdir(gconfig.summary)
    ))
    weeklist.sort(reverse=reverse)
    if single_week:
        return weeklist
    weeklist = list(filter(
        lambda x: os.path.exists(
            os.path.join(os.path.join(gconfig.summary, x),
                         f'{gconfig.prefix}项目工作周报-{x}.xlsx')
        ),
        weeklist
    ))
    return weeklist


async def swmg_group_list(week: str) -> list:
    source_dir = os.path.join(gconfig.summary, week)
    file_pattern = re.compile(f'{gconfig.prefix}小组工作周报-{week}-(.*?).xlsx')
    file_list = list(filter(
        lambda x: re.match(file_pattern, x),
        os.listdir(source_dir)
    ))
    file_list = list(map(
        lambda x: {
            'title': re.findall(file_pattern, x)[0],
            'key': x,
            'disabled': False
        },
        file_list
    ))
    disabled_group = list(filter(
        lambda x: x[3:] not in [z['title'] for z in file_list],
        gconfig.group
    ))
    for dis in disabled_group:
        file_list.append({
            'title': dis[3:],
            'key': dis,
            'disabled': True
        })
    return file_list


async def swmg_exec_merge(week: str, filelist: list, force: bool = False) -> tuple:
    source_dir = os.path.join(gconfig.summary, week)
    pattern = re.compile(f'{gconfig.prefix}小组工作周报-{week}-(.*?).xlsx')
    dist_path = os.path.join(source_dir, f'{gconfig.prefix}小组工作周报-{week}.xlsx')
    if not force and os.path.exists(dist_path):
        return False, dist_path
    try:
        await __exec(source_dir, filelist, dist_path, gconfig.template_project, pattern)
        return True, dist_path
    except Exception as ept:
        return False, f'{ept}'


async def swsg_name_list(source_dir: str, week: str) -> list:
    file_pattern = re.compile(f'{gconfig.prefix}个人工作周报-{week}-(.*?).xlsx')
    file_list = list(filter(
        lambda x: re.match(file_pattern, x),
        os.listdir(source_dir)
    ))
    file_list = list(map(
        lambda x: {
            'title': re.findall(file_pattern, x)[0],
            'key': x
        },
        file_list
    ))
    return file_list


async def swsg_exec_merge(group_name: str, week: str, filelist: list, force: bool = False) -> tuple:
    source_dir = os.path.join(
        gconfig.workspace, os.path.join(group_name, week)
    )
    pattern = re.compile(f'{gconfig.prefix}个人工作周报-{week}-(.*).xlsx')
    dist_dir = os.path.join(gconfig.summary, week)
    if not os.path.exists(dist_dir):
        os.makedirs(dist_dir)
    dist_path = os.path.join(
        dist_dir,
        f'{gconfig.prefix}小组工作周报-{week}-{group_name[3:]}.xlsx'
    )
    if not force and os.path.exists(dist_path):
        return False, dist_path
    try:
        await __exec(source_dir, filelist, dist_path, gconfig.template_group, pattern)
        return True, dist_path
    except Exception as ept:
        return False, f'{ept}'


def check_port_in_use(port, host='127.0.0.1'):
    s = None
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        s.settimeout(1)
        s.connect((host, int(port)))
        return True
    except socket.error:
        return False
    finally:
        if s:
            s.close()
